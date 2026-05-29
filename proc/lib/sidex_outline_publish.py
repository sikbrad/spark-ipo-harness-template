#!/usr/bin/env python3
"""Publish SIDEX 2026 visit notes to Outline."""

from __future__ import annotations

import argparse
import hashlib
import json
import mimetypes
import re
import time
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from PIL import Image, ImageOps


ROOT = Path("/Users/gq/works/projs/dof-work-skills/dof-work-startpoint-04")
ANALYSIS_JSON = ROOT / "output/sidex-2026/qwen-vqa/sidex_qwen_vendor_analysis.json"
FEEDBACK_XLSX = ROOT / "output/sidex-2026/qwen-vqa/SIDEX_2026_feedback_memo.xlsx"
PUBLISH_ROOT = ROOT / "output/sidex-2026/outline_publish"
IMAGE_ROOT = PUBLISH_ROOT / "images"
ASSET_CACHE = PUBLISH_ROOT / "outline_assets.json"
DOC_CACHE = PUBLISH_ROOT / "outline_docs.json"
PUBLISH_DATASET = PUBLISH_ROOT / "publish_dataset.json"
ROTATION_OVERRIDES = PUBLISH_ROOT / "rotation_overrides.json"

OUTLINE_BASE = "https://outline.doflab.com"
TARGET_URL_ID = "Pi7l3wBPSd"
REQUEST_TIMEOUT = 60
IMAGE_MAX_EDGE = 1600
IMAGE_QUALITY = 84

ROLE_LABEL = {
    "snaps": "현장 사진",
    "handouts": "자료 사진",
}


@dataclass
class OutlineDocument:
    id: str
    url_id: str
    title: str
    url: str


def read_env_key(name: str) -> str:
    env_path = ROOT / ".env"
    for line in env_path.read_text(encoding="utf-8").splitlines():
        if line.startswith(f"{name}="):
            return line.split("=", 1)[1].strip().strip("'\"")
    raise RuntimeError(f"{name} not found")


def compact(value: Any, sep: str = " / ") -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return sep.join(str(v).strip() for v in value if str(v).strip())
    return str(value).strip()


def md_escape(text: str) -> str:
    return (text or "").replace("|", "\\|").replace("\n", "<br>")


def slugish(text: str) -> str:
    value = re.sub(r"[^0-9A-Za-z가-힣]+", "-", text).strip("-")
    return value[:80] or "document"


def short_vendor_name(name: str) -> str:
    value = name or "미확인 업체"
    value = re.sub(r"\s+", " ", value).strip()
    return value


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_feedback() -> dict[str, dict[str, str]]:
    workbook = load_workbook(FEEDBACK_XLSX, data_only=False)
    sheet = workbook["업체별 피드백"]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    indices = {name: headers.index(name) + 1 for name in headers if name}
    required = ["업체명", "백인식메모"]
    missing = [name for name in required if name not in indices]
    if missing:
        raise RuntimeError(f"missing feedback columns: {missing}")
    feedback: dict[str, dict[str, str]] = {}
    for row in range(2, sheet.max_row + 1):
        vendor = sheet.cell(row, indices["업체명"]).value
        if not vendor:
            continue
        memo = sheet.cell(row, indices["백인식메모"]).value
        status = sheet.cell(row, indices.get("검토상태", 1)).value if "검토상태" in indices else ""
        feedback[str(vendor).strip()] = {
            "memo": str(memo).strip() if memo else "",
            "status": str(status).strip() if status else "",
        }
    return feedback


def all_photos(vendor: dict[str, Any]) -> list[dict[str, Any]]:
    return [*(vendor.get("snaps") or []), *(vendor.get("handouts") or [])]


def find_feedback_for_vendor(vendor: str, feedback: dict[str, dict[str, str]]) -> dict[str, str]:
    if vendor in feedback:
        return feedback[vendor]
    base = re.sub(r"\s+", "", vendor or "")
    for key, value in feedback.items():
        if re.sub(r"\s+", "", key) == base:
            return value
    return {"memo": "", "status": ""}


def rotate_image(image: Image.Image, degrees_clockwise: int) -> Image.Image:
    degrees_clockwise %= 360
    if degrees_clockwise == 0:
        return image
    return image.rotate(-degrees_clockwise, expand=True)


def prepare_image(source: Path, role: str, order: int, overrides: dict[str, int]) -> dict[str, Any]:
    IMAGE_ROOT.mkdir(parents=True, exist_ok=True)
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()[:16]
    rotation = int(overrides.get(source.name, overrides.get(str(source), 0)) or 0)
    output_name = f"{role}_{order:03d}_{source.stem}_r{rotation}_{source_hash}.jpg"
    output = IMAGE_ROOT / output_name

    with Image.open(source) as raw:
        image = ImageOps.exif_transpose(raw)
        original_size = list(raw.size)
        transposed_size = list(image.size)
        image = rotate_image(image, rotation)
        image = image.convert("RGB")
        image.thumbnail((IMAGE_MAX_EDGE, IMAGE_MAX_EDGE), Image.Resampling.LANCZOS)
        final_size = list(image.size)
        if not output.exists():
            image.save(output, "JPEG", quality=IMAGE_QUALITY, optimize=True)

    return {
        "source": str(source),
        "published_path": str(output),
        "published_name": output.name,
        "sha256": hashlib.sha256(output.read_bytes()).hexdigest(),
        "role": role,
        "order": order,
        "rotation_applied_clockwise": rotation,
        "original_size": original_size,
        "transposed_size": transposed_size,
        "published_size": final_size,
    }


class OutlineClient:
    def __init__(self, key: str):
        self.session = requests.Session()
        self.session.headers.update({"Authorization": f"Bearer {key}"})

    def api(self, endpoint: str, body: dict[str, Any]) -> dict[str, Any]:
        url = f"{OUTLINE_BASE}/api/{endpoint}"
        response = self.session.post(url, json=body, timeout=REQUEST_TIMEOUT)
        if response.status_code >= 400:
            raise RuntimeError(f"{endpoint} failed {response.status_code}: {response.text[:1000]}")
        return response.json()

    def document_info(self, doc_id: str) -> dict[str, Any]:
        return self.api("documents.info", {"id": doc_id})["data"]

    def documents_list(self, **params: Any) -> list[dict[str, Any]]:
        offset = 0
        docs: list[dict[str, Any]] = []
        while True:
            body = {"limit": 100, "offset": offset, **params}
            data = self.api("documents.list", body)
            batch = data.get("data") or []
            docs.extend(batch)
            pagination = data.get("pagination") or {}
            total = pagination.get("total", len(docs))
            offset += len(batch)
            if not batch or offset >= total:
                return docs

    def create_document(self, title: str, parent_document_id: str, text: str) -> dict[str, Any]:
        return self.api(
            "documents.create",
            {
                "title": title,
                "parentDocumentId": parent_document_id,
                "text": text,
                "publish": True,
                "fullWidth": True,
            },
        )["data"]

    def update_document(self, doc_id: str, title: str, text: str) -> dict[str, Any]:
        return self.api(
            "documents.update",
            {
                "id": doc_id,
                "title": title,
                "text": text,
                "publish": True,
                "fullWidth": True,
            },
        )["data"]

    def create_attachment(self, document_id: str, image_path: Path) -> dict[str, Any]:
        content_type = mimetypes.guess_type(image_path.name)[0] or "image/jpeg"
        create = self.api(
            "attachments.create",
            {
                "name": image_path.name,
                "documentId": document_id,
                "contentType": content_type,
                "size": image_path.stat().st_size,
            },
        )["data"]
        form = create.get("form") or {}
        upload_url = create["uploadUrl"]
        if upload_url.startswith("/"):
            upload_url = f"{OUTLINE_BASE}{upload_url}"
        with image_path.open("rb") as file_obj:
            upload_response = requests.post(
                upload_url,
                data=form,
                files={"file": (image_path.name, file_obj, content_type)},
                headers={"Authorization": self.session.headers["Authorization"]},
                timeout=REQUEST_TIMEOUT,
            )
        if upload_response.status_code >= 400:
            raise RuntimeError(f"attachment upload failed {upload_response.status_code}: {upload_response.text[:1000]}")
        return create["attachment"]


def doc_url(doc: dict[str, Any]) -> str:
    return f"{OUTLINE_BASE}/doc/{doc['urlId']}"


def build_vendor_summary(vendor: dict[str, Any], memo: str) -> str:
    pieces = []
    if memo:
        pieces.append(memo)
    if vendor.get("claims"):
        pieces.append("주요 메시지는 " + compact(vendor["claims"][:4]) + "입니다.")
    if vendor.get("product_names"):
        pieces.append("제품/브랜드로는 " + compact(vendor["product_names"][:6]) + "가 확인됩니다.")
    context = []
    for photo in all_photos(vendor)[:3]:
        if photo.get("visual_context"):
            context.append(photo["visual_context"])
    if context:
        pieces.append("사진상 맥락: " + compact(context, " "))
    return "\n\n".join(pieces).strip()


def build_child_markdown(vendor: dict[str, Any], memo: str, image_urls: dict[str, str]) -> str:
    photos = all_photos(vendor)
    lines = []
    if memo:
        lines += ["## 현장 인사이트", "", memo, ""]
    lines += ["## 한눈에 보기", ""]
    if vendor.get("booth"):
        lines.append(f"- 부스: {vendor['booth']}")
    sectors = compact(vendor.get("sectors") or [], ", ")
    if sectors:
        lines.append(f"- 분야: {sectors}")
    products = compact(vendor.get("product_names") or [], ", ")
    if products:
        lines.append(f"- 제품/브랜드: {products}")
    if vendor.get("claims"):
        lines.append("- 회사가 강조한 점:")
        for claim in vendor["claims"][:8]:
            lines.append(f"  - {claim}")
    lines.append("")

    visible_text = []
    for photo in photos:
        for text in photo.get("visible_text") or []:
            if text and text not in visible_text:
                visible_text.append(text)
    if visible_text:
        lines += ["## 현장에서 확인된 문구", ""]
        for text in visible_text[:24]:
            lines.append(f"- {text}")
        lines.append("")

    summary = build_vendor_summary(vendor, memo)
    if summary:
        lines += ["## 정리", "", summary, ""]

    if photos:
        lines += ["## 사진", ""]
        for photo in photos:
            source = photo.get("source", "")
            url = image_urls.get(source)
            if not url:
                continue
            label = ROLE_LABEL.get(photo.get("role"), "사진")
            caption = f"{label} · {photo.get('timestamp') or ''} · {photo.get('filename') or ''}".strip(" ·")
            lines += [f"![{caption}]({url})", "", caption, ""]

    return "\n".join(lines).strip() + "\n"


def sector_counts(vendors: list[dict[str, Any]]) -> list[tuple[str, int]]:
    counter: Counter[str] = Counter()
    for vendor in vendors:
        for sector in vendor.get("sectors") or []:
            counter[sector] += 1
    return counter.most_common()


def build_parent_markdown(
    vendors: list[dict[str, Any]],
    docs: dict[str, OutlineDocument],
    feedback: dict[str, dict[str, str]],
) -> str:
    counts = {
        "vendors": len(vendors),
        "photos": sum(len(all_photos(v)) for v in vendors),
        "memo": sum(1 for v in vendors if find_feedback_for_vendor(v.get("vendor", ""), feedback).get("memo")),
    }
    memo_vendors = []
    for vendor in vendors:
        memo = find_feedback_for_vendor(vendor.get("vendor", ""), feedback).get("memo", "")
        if memo:
            memo_vendors.append((vendor, memo))

    lines = [
        "# SIDEX 2026 방문 요약",
        "",
        "## 핵심 현장 인사이트",
        "",
    ]
    if memo_vendors:
        for vendor, memo in memo_vendors:
            title = short_vendor_name(vendor.get("vendor", ""))
            booth = f" / {vendor.get('booth')}" if vendor.get("booth") else ""
            lines.append(f"- **{title}{booth}**: {memo}")
    else:
        lines.append("- 현장 메모가 입력된 업체가 없습니다.")

    lines += [
        "",
        "## 방문 기록 개요",
        "",
        f"- 정리 업체: {counts['vendors']}개",
        f"- 정리 사진: {counts['photos']}장",
        f"- 현장 메모 반영 업체: {counts['memo']}개",
        "",
        "## 분야별 분포",
        "",
        "| 분야 | 업체 수 |",
        "| --- | ---: |",
    ]
    for sector, count in sector_counts(vendors):
        lines.append(f"| {md_escape(sector)} | {count} |")

    lines += [
        "",
        "## 전시장에서 읽힌 큰 흐름",
        "",
        "- **디지털 워크플로우의 웹 전환**: 케이스 공유, 상담, 클라우드 기반 뷰어처럼 링크 중심으로 협업하려는 흐름이 확인됩니다.",
        "- **저가형 장비 경쟁 심화**: 중국 장비와 OEM, 렌탈/차감 계약을 결합한 밀링기·스캐너·퍼네스·3D 프린터 패키지가 눈에 띕니다.",
        "- **대형 임플란트사의 플랫폼화**: 임플란트, 장비, 진단, 이벤트, 소모품까지 수직계열화한 대형 부스가 강한 존재감을 보였습니다.",
        "- **스캐너 시장의 다층화**: 고성능/브랜드 제품과 저가형 OEM 제품, 모바일 연동 제품이 동시에 경쟁하고 있습니다.",
        "- **치과 주변 서비스 확장**: 로봇, 상담 솔루션, 미용 장비, 산업디자인 등 진료 주변 영역의 제품화가 늘고 있습니다.",
        "",
        "## 업체별 문서",
        "",
        "| 업체 | 부스 | 분야 | 사진 | 메모 | 문서 |",
        "| --- | --- | --- | ---: | --- | --- |",
    ]
    for vendor in vendors:
        name = short_vendor_name(vendor.get("vendor", ""))
        doc = docs.get(name)
        memo = find_feedback_for_vendor(vendor.get("vendor", ""), feedback).get("memo", "")
        link = f"[열기]({doc.url})" if doc else ""
        lines.append(
            "| "
            + " | ".join(
                [
                    md_escape(name),
                    md_escape(vendor.get("booth", "")),
                    md_escape(compact(vendor.get("sectors") or [], ", ")),
                    str(len(all_photos(vendor))),
                    "있음" if memo else "",
                    link,
                ]
            )
            + " |"
        )
    return "\n".join(lines).strip() + "\n"


def build_dataset() -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], dict[str, Any]]:
    PUBLISH_ROOT.mkdir(parents=True, exist_ok=True)
    analysis = load_json(ANALYSIS_JSON, {})
    feedback = load_feedback()
    overrides = load_json(ROTATION_OVERRIDES, {})
    vendors = analysis.get("vendors") or []
    for vendor in vendors:
        vendor_name = vendor.get("vendor") or "미확인 업체"
        feedback_item = find_feedback_for_vendor(vendor_name, feedback)
        vendor["feedback_memo"] = feedback_item.get("memo", "")
        vendor["feedback_status"] = feedback_item.get("status", "")
        for photo in all_photos(vendor):
            source = Path(photo["source"])
            prepared = prepare_image(source, photo.get("role") or "photo", int(photo.get("path_order") or 0), overrides)
            photo["outline_image"] = prepared
    dataset = {
        "counts": analysis.get("counts") or {},
        "vendors": vendors,
        "feedback_nonempty": {
            vendor.get("vendor") or "미확인 업체": vendor.get("feedback_memo")
            for vendor in vendors
            if vendor.get("feedback_memo")
        },
    }
    save_json(PUBLISH_DATASET, dataset)
    return vendors, feedback, dataset


def upload_vendor_images(
    client: OutlineClient,
    document_id: str,
    vendor: dict[str, Any],
    asset_cache: dict[str, Any],
) -> dict[str, str]:
    urls: dict[str, str] = {}
    for photo in all_photos(vendor):
        image = photo.get("outline_image") or {}
        source = image.get("source")
        published_path = Path(image.get("published_path", ""))
        if not source or not published_path.exists():
            continue
        cache_key = f"{document_id}:{source}:{image.get('sha256')}"
        cached = asset_cache.get(cache_key)
        if cached and cached.get("url"):
            urls[source] = cached["url"]
            continue
        attachment = client.create_attachment(document_id, published_path)
        asset_cache[cache_key] = {
            "source": source,
            "published_path": str(published_path),
            "document_id": document_id,
            "url": attachment.get("url", ""),
            "name": attachment.get("name", published_path.name),
        }
        urls[source] = attachment.get("url", "")
        save_json(ASSET_CACHE, asset_cache)
        time.sleep(0.05)
    return urls


def publish(dry_run: bool = False, limit: int | None = None) -> dict[str, Any]:
    vendors, feedback, dataset = build_dataset()
    if limit:
        vendors = vendors[:limit]
    key = read_env_key("DOF_OUTLINE_KEY")
    client = OutlineClient(key)
    parent = client.document_info(TARGET_URL_ID)
    existing_children = {
        doc["title"]: doc
        for doc in client.documents_list(parentDocumentId=parent["id"])
        if doc.get("title")
    }
    asset_cache = load_json(ASSET_CACHE, {})
    doc_cache = load_json(DOC_CACHE, {})
    published_docs: dict[str, OutlineDocument] = {}

    if dry_run:
        return {"dry_run": True, "vendors": len(vendors), "parent": parent, "dataset": dataset}

    for index, vendor in enumerate(vendors, start=1):
        vendor_name = short_vendor_name(vendor.get("vendor", ""))
        title = f"{index:02d}. {vendor_name}"
        if vendor.get("booth"):
            title += f" / {vendor['booth']}"

        doc = existing_children.get(title)
        if not doc:
            doc = client.create_document(title, parent["id"], "작성 중입니다.")
            time.sleep(0.1)

        image_urls = upload_vendor_images(client, doc["id"], vendor, asset_cache)
        body = build_child_markdown(vendor, vendor.get("feedback_memo", ""), image_urls)
        updated = client.update_document(doc["id"], title, body)
        child = OutlineDocument(id=updated["id"], url_id=updated["urlId"], title=updated["title"], url=doc_url(updated))
        published_docs[vendor_name] = child
        doc_cache[vendor_name] = child.__dict__
        save_json(DOC_CACHE, doc_cache)
        print(json.dumps({"published": title, "images": len(image_urls), "url": child.url}, ensure_ascii=False))
        time.sleep(0.1)

    parent_title = "SIDEX 2026 방문 요약"
    parent_body = build_parent_markdown(vendors, published_docs, feedback)
    updated_parent = client.update_document(parent["id"], parent_title, parent_body)
    result = {
        "parent": {
            "id": updated_parent["id"],
            "urlId": updated_parent["urlId"],
            "title": updated_parent["title"],
            "url": doc_url(updated_parent),
        },
        "children": [doc.__dict__ for doc in published_docs.values()],
        "counts": {
            "vendors": len(vendors),
            "photos": sum(len(all_photos(v)) for v in vendors),
            "memos": sum(1 for v in vendors if v.get("feedback_memo")),
        },
    }
    save_json(PUBLISH_ROOT / "publish_result.json", result)
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()
    result = publish(dry_run=args.dry_run, limit=args.limit)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
