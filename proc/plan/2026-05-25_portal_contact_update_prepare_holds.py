#!/usr/bin/env python3
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path("/Users/gq/works/projs/dof-work-skills/dof-work-startpoint-04")
OUT_DIR = BASE_DIR / "output/portal-contact-update-20260525"
HOLDS_XLSX = OUT_DIR / "contact-update-holds.xlsx"
ANALYSIS_JSON = OUT_DIR / "contact-update-analysis.json"
MIGRATION_SQL = Path(
    "/Users/gq/works/projs/dofing-order-app/order-web/dof-order-app/dof-order-web-3-az/apps/server/prisma/migrations/008_company_contact_fields.sql"
)
OUTPUT_PROD_SQL = OUT_DIR / "contact-update-holds-prod.sql"
OUTPUT_PROD_JSON = OUT_DIR / "contact-update-holds-apply-final-prod.json"


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    text = str(value)
    return "'" + text.replace("'", "''") + "'"


def sql_value(value: str) -> str:
    return "NULL" if value == "" else sql_literal(value)


def load_hold_decisions() -> list[dict[str, Any]]:
    wb = load_workbook(HOLDS_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = {clean(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)}
    required = ["excelRow", "결정", "적용할 company id", "메모"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f"missing hold decision columns: {missing}")

    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        decision = clean(ws.cell(row=r, column=headers["결정"]).value).lower()
        if decision != "apply":
            continue
        excel_row = int(ws.cell(row=r, column=headers["excelRow"]).value)
        company_id = int(ws.cell(row=r, column=headers["적용할 company id"]).value)
        rows.append(
            {
                "excelRow": excel_row,
                "companyId": company_id,
                "decisionMemo": clean(ws.cell(row=r, column=headers["메모"]).value),
            }
        )
    return rows


def build_updates() -> list[dict[str, Any]]:
    analysis = {int(row["excelRow"]): row for row in json.loads(ANALYSIS_JSON.read_text(encoding="utf-8"))}
    updates: list[dict[str, Any]] = []
    for decision in load_hold_decisions():
        source = analysis.get(decision["excelRow"])
        if not source:
            raise RuntimeError(f"analysis row not found: {decision['excelRow']}")

        target_office = clean(source.get("targetOfficePhone"))
        target_mobile1 = clean(source.get("targetMobile1"))
        target_mobile2 = clean(source.get("targetMobile2"))
        target_email1 = clean(source.get("targetEmail1")).lower()
        target_email2 = clean(source.get("targetEmail2")).lower()

        set_values: dict[str, str | None] = {}
        if target_office:
            set_values["phone"] = target_office
        elif target_mobile1:
            set_values["phone"] = None

        if target_email1:
            set_values["email"] = target_email1
        if target_email2:
            set_values["email2"] = target_email2
        if target_mobile1:
            set_values["mobile1"] = target_mobile1
        if target_mobile2:
            set_values["mobile2"] = target_mobile2

        if not set_values:
            continue

        updates.append(
            {
                **decision,
                "fgName": clean(source.get("fgName")),
                "taxName": clean(source.get("taxName")),
                "targetOfficePhone": target_office,
                "targetMobile1": target_mobile1,
                "targetMobile2": target_mobile2,
                "targetEmail1": target_email1,
                "targetEmail2": target_email2,
                "setValues": set_values,
            }
        )
    return updates


def contact_memo_sql(excel_row: int, fg_name: str) -> str:
    marker = f"통계 row {excel_row}"
    origin = f"원본: 통계 row {excel_row} / {fg_name or '-'}"
    return (
        "CASE "
        f'WHEN "contactMemo" IS NOT NULL AND position({sql_literal(marker)} in "contactMemo") > 0 THEN "contactMemo" '
        "ELSE concat_ws(E'\\n\\n', NULLIF(\"contactMemo\", ''), "
        "'---- AX ----' || E'\\n' || "
        f"{sql_literal(origin)} || E'\\n' || "
        "'업체명: ' || COALESCE(\"name\", '') || E'\\n' || "
        "'상호: ' || COALESCE(\"businessName\", '') || E'\\n' || "
        "'ERP ID: ' || COALESCE(\"erpId\", '') || E'\\n' || "
        "'유선번호: ' || COALESCE(\"phone\", '') || E'\\n' || "
        "'팩스번호: ' || COALESCE(\"fax\", '') || E'\\n' || "
        "'이메일1: ' || COALESCE(\"email\", '') || E'\\n' || "
        "'이메일2: ' || COALESCE(\"email2\", '') || E'\\n' || "
        "'웹사이트: ' || COALESCE(\"website\", '') || E'\\n' || "
        "'모바일1: ' || COALESCE(\"mobile1\", '') || E'\\n' || "
        "'모바일2: ' || COALESCE(\"mobile2\", '') || E'\\n' || "
        "'------------') END"
    )


def write_sql(path: Path, updates: list[dict[str, Any]], target: str) -> None:
    statements = [
        "-- Generated by proc/plan/2026-05-25_portal_contact_update_prepare_holds.py",
        f"-- Generated at {datetime.now().isoformat(timespec='seconds')}",
        f"-- Target: {target}",
        f"-- Hold updates: {len(updates)}",
        "BEGIN;",
        MIGRATION_SQL.read_text(encoding="utf-8").strip(),
    ]

    for row in updates:
        set_parts = []
        for field, value in row["setValues"].items():
            set_parts.append(f'"{field}" = {sql_value(value) if value is not None else "NULL"}')
        set_parts.append(f'"contactMemo" = {contact_memo_sql(row["excelRow"], row["fgName"])}')
        set_parts.append('"updatedAt" = NOW()')
        statements.append(
            f"-- 통계 row {row['excelRow']} / {row['fgName']} / company id {row['companyId']}\n"
            "UPDATE \"Company\"\n"
            "SET " + ",\n    ".join(set_parts) + "\n"
            f"WHERE id = {int(row['companyId'])} AND \"deletedAt\" IS NULL;"
        )

    statements.append("COMMIT;")
    path.write_text("\n\n".join(statements) + "\n", encoding="utf-8")


def main() -> None:
    updates = build_updates()
    OUTPUT_PROD_JSON.write_text(json.dumps(updates, ensure_ascii=False, indent=2), encoding="utf-8")
    write_sql(OUTPUT_PROD_SQL, updates, "prod")
    print(f"updates={len(updates)}")
    print(OUTPUT_PROD_JSON)
    print(OUTPUT_PROD_SQL)


if __name__ == "__main__":
    main()
