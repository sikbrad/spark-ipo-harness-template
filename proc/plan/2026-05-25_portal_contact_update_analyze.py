#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


XLSX_PATH = Path(
    "/Users/gq/works/projs/crm-migration/dofing_crm_facade/output/pack/latest_data/contact/customer_contacts_latest.xlsx"
)
SHEET_NAME = "통계"
DEFAULT_OUTPUT_DIR = Path(
    "/Users/gq/works/projs/dof-work-skills/dof-work-startpoint-04/output/portal-contact-update-20260525"
)
PSQL = "/opt/homebrew/opt/libpq/bin/psql"


EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_name(value: str) -> str:
    value = clean_text(value).lower()
    value = re.sub(r"\[[^\]]+\]", "", value)
    value = re.sub(r"\(구\)", "", value)
    value = re.sub(r"[\s\u3000·.,'/\\\-_[\]{}]", "", value)
    value = value.replace("㈜", "주").replace("(주)", "주").replace("（주）", "주")
    value = value.replace("주식회사", "주")
    return value


def normalize_name_loose(value: str) -> str:
    value = normalize_name(value)
    value = re.sub(r"\([^)]*\)", "", value)
    value = re.sub(r"(치과의원|치과기공소|기공소|덴탈랩|덴탈|주)$", "", value)
    return value


def normalize_digits(value: Any) -> str:
    raw = clean_text(value)
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 10 and digits.startswith("10"):
        digits = "0" + digits
    return digits


def format_phone(digits: str) -> str:
    if not digits:
        return ""
    if digits.startswith("010") and len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if digits.startswith("02"):
        if len(digits) == 9:
            return f"{digits[:2]}-{digits[2:5]}-{digits[5:]}"
        if len(digits) == 10:
            return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return digits


def split_emails(*values: Any) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        for email in EMAIL_RE.findall(clean_text(value)):
            key = email.lower()
            if key not in seen:
                seen.add(key)
                result.append(key)
    return result


def split_phones(*values: Any) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        raw = clean_text(value)
        if not raw:
            continue
        parts = re.split(r"[,;/\n]|(?:\s{2,})", raw)
        for part in parts:
            digits = normalize_digits(part)
            if not digits or len(digits) < 8:
                continue
            if digits not in seen:
                seen.add(digits)
                result.append(format_phone(digits))
    return result


def is_mobile(phone: str) -> bool:
    return normalize_digits(phone).startswith("010")


def psql_json(database_url: str, sql: str) -> Any:
    wrapped = "SELECT COALESCE((" + sql.rstrip().rstrip(";") + "), '[]'::json)::text"
    proc = subprocess.run(
        [PSQL, database_url, "-X", "-v", "ON_ERROR_STOP=1", "-Aqt", "-c", wrapped],
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    text = proc.stdout.strip()
    return json.loads(text) if text else None


def load_contacts() -> list[dict[str, Any]]:
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    headers = [clean_text(cell) for cell in rows[0]]
    contacts = []
    for excel_row, row in enumerate(rows[1:], start=2):
        item = {headers[i] or f"col_{i + 1}": row[i] if i < len(row) else None for i in range(len(headers))}
        emails = split_emails(item.get("이메일"), item.get("이메일1"))
        phones = split_phones(item.get("전화번호"), item.get("번호1"), item.get("번호2"))
        mobile_numbers = [p for p in phones if is_mobile(p)]
        office_numbers = [p for p in phones if not is_mobile(p)]
        names = [
            clean_text(item.get("회사명 (FG 기준)")),
            clean_text(item.get("회사명(세금계산서기준)")),
            clean_text(item.get("다른이름1")),
        ]
        names = [name for name in names if name]
        contacts.append(
            {
                "excelRow": excel_row,
                "seq": clean_text(item.get("서순")),
                "duplicate": clean_text(item.get("중복")),
                "erpId": clean_text(item.get("ERP ID")),
                "fgName": clean_text(item.get("회사명 (FG 기준)")),
                "taxName": clean_text(item.get("회사명(세금계산서기준)")),
                "altName1": clean_text(item.get("다른이름1")),
                "phoneRaw": clean_text(item.get("전화번호")),
                "emailRaw": clean_text(item.get("이메일")),
                "email1Raw": clean_text(item.get("이메일1")),
                "phone1Raw": clean_text(item.get("번호1")),
                "phone2Raw": clean_text(item.get("번호2")),
                "sendTarget": clean_text(item.get("전송대상")),
                "issue2512": clean_text(item.get("문제사유2512")),
                "issue2601": clean_text(item.get("문제사유2601")),
                "issue2601Feedback": clean_text(item.get("문제사유2601피드백")),
                "issue2602": clean_text(item.get("문제사유2602")),
                "candidateNames": names,
                "targetEmail1": emails[0] if len(emails) > 0 else "",
                "targetEmail2": emails[1] if len(emails) > 1 else "",
                "allEmails": emails,
                "targetOfficePhone": office_numbers[0] if len(office_numbers) > 0 else "",
                "targetMobile1": mobile_numbers[0] if len(mobile_numbers) > 0 else "",
                "targetMobile2": mobile_numbers[1] if len(mobile_numbers) > 1 else "",
                "allPhones": phones,
                "allMobilePhones": mobile_numbers,
                "allOfficePhones": office_numbers,
            }
        )
    return contacts


def load_companies(database_url: str) -> list[dict[str, Any]]:
    sql = """
    SELECT json_agg(row_to_json(t))
    FROM (
      SELECT
        c.id,
        c.uk,
        c.name,
        c."businessName",
        c."erpId",
        c."businessNo",
        c.phone,
        c.fax,
        c.website,
        c.email,
        c.memo,
        c."importSource",
        c."deletedAt",
        COALESCE(
          array_agg(h.tag ORDER BY h.tag)
          FILTER (WHERE h.id IS NOT NULL AND h.type = 'alias' AND h."deletedAt" IS NULL),
          ARRAY[]::text[]
        ) AS aliases
      FROM "Company" c
      LEFT JOIN "Hashtag" h
        ON h."entityType" = 'COMPANY'
       AND h."entityUk" = c.uk
       AND h."companyUk" = c.uk
       AND h.type = 'alias'
       AND h."deletedAt" IS NULL
      WHERE c."deletedAt" IS NULL
      GROUP BY c.id
      ORDER BY c.name, c.id
    ) t
    """
    return psql_json(database_url, sql) or []


@dataclass
class Match:
    status: str
    tier: str
    company: dict[str, Any] | None
    candidates: list[dict[str, Any]]
    note: str


def build_indexes(companies: list[dict[str, Any]]) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    strict: dict[str, list[dict[str, Any]]] = defaultdict(list)
    loose: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for company in companies:
        names = [company.get("name"), company.get("businessName"), *(company.get("aliases") or [])]
        for name in names:
            n1 = normalize_name(clean_text(name))
            n2 = normalize_name_loose(clean_text(name))
            if n1:
                strict[n1].append(company)
            if n2:
                loose[n2].append(company)
    return strict, loose


def unique(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    by_uk = {c["uk"]: c for c in candidates}
    return next(iter(by_uk.values())) if len(by_uk) == 1 else None


def find_match(contact: dict[str, Any], companies: list[dict[str, Any]], strict: dict[str, list[dict[str, Any]]], loose: dict[str, list[dict[str, Any]]]) -> Match:
    erp_id = contact["erpId"]
    if erp_id:
        erp_matches = [c for c in companies if clean_text(c.get("erpId")) == erp_id]
        found = unique(erp_matches)
        if found:
            return Match("matched", "erpId", found, erp_matches, "ERP ID exact match")
        if erp_matches:
            return Match("ambiguous", "erpId", None, erp_matches, "ERP ID matched multiple companies")

    for name in contact["candidateNames"]:
        key = normalize_name(name)
        matches = strict.get(key, [])
        found = unique(matches)
        if found:
            return Match("matched", "strict-name", found, matches, f"exact normalized name: {name}")
        if matches:
            return Match("ambiguous", "strict-name", None, matches, f"name matched multiple companies: {name}")

    for name in contact["candidateNames"]:
        key = normalize_name_loose(name)
        matches = loose.get(key, [])
        found = unique(matches)
        if found:
            return Match("matched", "loose-name", found, matches, f"loose normalized name: {name}")
        if matches:
            return Match("ambiguous", "loose-name", None, matches, f"loose name matched multiple companies: {name}")

    best: list[tuple[float, dict[str, Any], str]] = []
    contact_keys = [normalize_name(name) for name in contact["candidateNames"] if normalize_name(name)]
    for company in companies:
        company_names = [company.get("name"), company.get("businessName"), *(company.get("aliases") or [])]
        for cname in company_names:
            ckey = normalize_name(clean_text(cname))
            if not ckey:
                continue
            score = max((SequenceMatcher(None, q, ckey).ratio() for q in contact_keys), default=0)
            if score >= 0.88:
                best.append((score, company, clean_text(cname)))
    best.sort(key=lambda x: (-x[0], x[1]["name"]))
    compact: list[dict[str, Any]] = []
    seen: set[str] = set()
    for score, company, matched_name in best:
        if company["uk"] in seen:
            continue
        seen.add(company["uk"])
        item = dict(company)
        item["_score"] = round(score, 4)
        item["_matchedName"] = matched_name
        compact.append(item)
        if len(compact) >= 5:
            break
    if len(compact) == 1 and compact[0]["_score"] >= 0.965:
        return Match("review", "fuzzy-high", compact[0], compact, "single high-score fuzzy candidate; manual review before update")
    if compact:
        return Match("unmatched", "fuzzy-candidates", None, compact, "no exact match; fuzzy candidates listed")
    return Match("unmatched", "none", None, [], "no company candidate found")


def normalize_contact_current(company: dict[str, Any] | None) -> dict[str, str]:
    if not company:
        return {}
    return {
        "currentPhone": format_phone(normalize_digits(company.get("phone"))),
        "currentEmail": clean_text(company.get("email")).lower(),
        "currentFax": format_phone(normalize_digits(company.get("fax"))),
        "currentWebsite": clean_text(company.get("website")),
        "currentMemo": clean_text(company.get("memo")),
    }


def compute_action(contact: dict[str, Any], match: Match) -> tuple[str, list[str]]:
    if match.status == "ambiguous":
        return "hold_ambiguous_match", ["ambiguous company match"]
    if match.status == "unmatched":
        return "hold_unmatched", ["company not found"]
    if match.status == "review":
        return "hold_review_fuzzy", ["fuzzy match requires manual approval"]
    assert match.company is not None
    cur = normalize_contact_current(match.company)
    reasons = []
    target_company_phone = contact["targetOfficePhone"] or contact["targetMobile1"]
    if target_company_phone and cur.get("currentPhone") != target_company_phone:
        reasons.append("phone")
    if contact["targetEmail1"] and cur.get("currentEmail") != contact["targetEmail1"]:
        reasons.append("email1")
    if contact["targetEmail2"]:
        reasons.append("email2-new-column")
    if contact["targetMobile1"]:
        reasons.append("mobile1-new-column")
    if contact["targetMobile2"]:
        reasons.append("mobile2-new-column")
    if reasons:
        return "update_needed", reasons
    return "already_latest", []


def previous_info_block(company: dict[str, Any]) -> str:
    rows = [
        f"업체명: {clean_text(company.get('name'))}",
        f"상호: {clean_text(company.get('businessName'))}",
        f"ERP ID: {clean_text(company.get('erpId'))}",
        f"유선/대표번호: {clean_text(company.get('phone'))}",
        f"팩스번호: {clean_text(company.get('fax'))}",
        f"이메일1: {clean_text(company.get('email'))}",
        "이메일2: ",
        f"웹사이트: {clean_text(company.get('website'))}",
        "모바일1: ",
        "모바일2: ",
    ]
    return "---- AX ----\n" + "\n".join(rows) + "\n------------"


def write_outputs(out_dir: Path, contacts: list[dict[str, Any]], companies: list[dict[str, Any]], matches: list[dict[str, Any]]) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    with (out_dir / "contact-update-analysis.json").open("w", encoding="utf-8") as f:
        json.dump(matches, f, ensure_ascii=False, indent=2)

    columns = [
        "excelRow",
        "seq",
        "fgName",
        "taxName",
        "altName1",
        "matchStatus",
        "matchTier",
        "action",
        "actionReasons",
        "companyId",
        "companyUk",
        "companyName",
        "companyBusinessName",
        "companyErpId",
        "currentPhone",
        "currentEmail",
        "currentFax",
        "currentWebsite",
        "targetOfficePhone",
        "targetMobile1",
        "targetMobile2",
        "targetEmail1",
        "targetEmail2",
        "candidateCount",
        "candidateSummary",
        "matchNote",
        "phoneRaw",
        "phone1Raw",
        "phone2Raw",
        "emailRaw",
        "email1Raw",
        "sendTarget",
        "issue2512",
        "issue2601",
        "issue2601Feedback",
        "issue2602",
    ]
    with (out_dir / "contact-update-analysis.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(matches)

    wb = Workbook()
    ws = wb.active
    ws.title = "analysis"
    ws.append(columns)
    for item in matches:
        ws.append([item.get(col, "") for col in columns])
    for col in ws.columns:
        max_len = max(len(clean_text(cell.value)) for cell in col[:200])
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)
    wb.save(out_dir / "contact-update-analysis.xlsx")

    summary = Counter(item["action"] for item in matches)
    match_summary = Counter(item["matchStatus"] for item in matches)
    reason_summary = Counter(reason for item in matches for reason in item["actionReasonsList"])
    dup_names = Counter((item["fgName"], item["taxName"]) for item in matches)
    duplicate_input_pairs = sum(1 for _key, count in dup_names.items() if count > 1)
    lines = [
        "# Portal contact update analysis",
        "",
        f"- Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"- Input workbook: `{XLSX_PATH}`",
        f"- Sheet: `{SHEET_NAME}`",
        f"- Input rows: {len(contacts)}",
        f"- Active prod companies loaded: {len(companies)}",
        f"- Duplicate input name pairs: {duplicate_input_pairs}",
        "",
        "## Action summary",
        "",
    ]
    for key, count in summary.most_common():
        lines.append(f"- {key}: {count}")
    lines += ["", "## Match summary", ""]
    for key, count in match_summary.most_common():
        lines.append(f"- {key}: {count}")
    lines += ["", "## Change reason summary", ""]
    for key, count in reason_summary.most_common():
        lines.append(f"- {key}: {count}")
    lines += [
        "",
        "## Files",
        "",
        "- `contact-update-analysis.xlsx`: manual review workbook",
        "- `contact-update-analysis.csv`: same data for grep/diff",
        "- `contact-update-analysis.json`: complete structured data",
    ]
    (out_dir / "summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database-url", required=True)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    contacts = load_contacts()
    companies = load_companies(args.database_url)
    strict, loose = build_indexes(companies)

    rows = []
    for contact in contacts:
        match = find_match(contact, companies, strict, loose)
        action, action_reasons = compute_action(contact, match)
        company = match.company
        cur = normalize_contact_current(company)
        candidates = match.candidates
        rows.append(
            {
                **contact,
                **cur,
                "matchStatus": match.status,
                "matchTier": match.tier,
                "action": action,
                "actionReasons": ", ".join(action_reasons),
                "actionReasonsList": action_reasons,
                "companyId": company.get("id") if company else "",
                "companyUk": company.get("uk") if company else "",
                "companyName": company.get("name") if company else "",
                "companyBusinessName": company.get("businessName") if company else "",
                "companyErpId": company.get("erpId") if company else "",
                "candidateCount": len(candidates),
                "candidateSummary": " | ".join(
                    f"{c.get('id')}:{c.get('name')}:{c.get('businessName') or ''}:{c.get('_score', '')}"
                    for c in candidates
                ),
                "matchNote": match.note,
                "previousInfoBlock": previous_info_block(company) if company else "",
            }
        )

    write_outputs(args.out_dir, contacts, companies, rows)
    print((args.out_dir / "summary.md").read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
